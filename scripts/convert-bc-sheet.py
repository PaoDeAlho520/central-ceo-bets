import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def unique_headers(row):
    headers = []
    seen = {}
    for index, value in enumerate(row, start=1):
        header = cell_text(value) or f"Coluna {index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        headers.append(header if count == 0 else f"{header} {count + 1}")
    return headers


def worksheet_rows(sheet):
    rows = [[cell_text(value) for value in row] for row in sheet.iter_rows(values_only=True)]
    while rows and not any(rows[-1]):
        rows.pop()
    while rows and not any(rows[0]):
        rows.pop(0)
    return rows


def find_header_index(rows):
    for index, row in enumerate(rows[:20]):
        filled = [value for value in row if value]
        if len(filled) >= 2:
            return index
    return 0


def records_from_rows(rows):
    if not rows:
        return [], []
    header_index = find_header_index(rows)
    headers = unique_headers(rows[header_index])
    records = []
    for row in rows[header_index + 1:]:
        if not any(row):
            continue
        padded = row + [""] * (len(headers) - len(row))
        records.append({header: padded[index] for index, header in enumerate(headers)})
    return headers, records


def convert(source_path, output_path):
    source = Path(source_path)
    output = Path(output_path)
    workbook = load_workbook(source, data_only=True, read_only=True)

    sheets = []
    row_count = 0
    for sheet in workbook.worksheets:
        rows = worksheet_rows(sheet)
        headers, records = records_from_rows(rows)
        if not headers and not records:
            continue
        row_count += len(records)
        sheets.append({
            "name": sheet.title,
            "headers": headers,
            "records": records,
            "recordCount": len(records),
        })

    payload = {
        "sourceFile": str(source),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sheetCount": len(sheets),
        "rowCount": row_count,
        "sheets": sheets,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"output": str(output), "sheetCount": len(sheets), "rowCount": row_count}, ensure_ascii=False))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python scripts/convert-bc-sheet.py origem.xlsx saida.json")
    convert(sys.argv[1], sys.argv[2])
