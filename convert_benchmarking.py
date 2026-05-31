import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def unique_headers(row):
    headers = []
    seen = {}
    for index, value in enumerate(row, start=1):
        header = cell_text(value) or f"Coluna {index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        headers.append(header if count == 0 else f"{header} {count + 1}")
    return headers


def worksheet_rows(sheet):
    rows = [
        [cell_text(value) for value in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    while rows and not any(rows[-1]):
        rows.pop()
    return rows


def records_from_rows(rows):
    if not rows:
        return [], [], 0
    header_index = find_header_index(rows)
    headers = unique_headers(rows[header_index])
    records = []
    for row in rows[header_index + 1:]:
        if not any(row):
            continue
        padded = row + [""] * (len(headers) - len(row))
        records.append({header: padded[index] for index, header in enumerate(headers)})
    return headers, records, header_index


def find_header_index(rows):
    signatures = [
        {"Country ID", "Country", "Brands"},
        {"#", "Brand", "APS"},
    ]
    for index, row in enumerate(rows):
        values = {cell_text(value) for value in row if cell_text(value)}
        if any(signature.issubset(values) for signature in signatures):
            return index
    return 0


def metadata_from_rows(rows, header_index):
    metadata = {}
    for row in rows[:header_index]:
        if len(row) >= 2 and row[0] and row[1]:
            metadata[row[0]] = row[1]
    return metadata


def sheet_metadata(name, records, rows, header_index):
    if name == "Ranking":
        return {"Type": "Country ranking"}
    metadata = metadata_from_rows(rows, header_index)
    if metadata:
        return metadata
    first = records[0] if records else {}
    return {
        "Country": name,
        "Top brand": first.get("Brand", ""),
        "Top APS": first.get("APS", ""),
        "Top CEB (US$)": first.get("CEB (US$)", ""),
    }


def convert(source_path, output_path):
    source = Path(source_path)
    output = Path(output_path)
    workbook = load_workbook(source, data_only=True, read_only=True)

    sheets = []
    row_count = 0
    ranking = []
    countries = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = worksheet_rows(sheet)
        headers, records, header_index = records_from_rows(rows)
        row_count += len(records)

        item = {
            "name": sheet_name,
            "headers": headers,
            "records": records,
            "raw": rows,
            "metadata": sheet_metadata(sheet_name, records, rows, header_index),
        }
        sheets.append(item)

        if sheet_name == "Ranking":
            ranking = records
        else:
            countries.append({
                "name": sheet_name,
                "records": records,
                "recordCount": len(records),
            })

    payload = {
        "sourceFile": str(source),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sheetCount": len(sheets),
        "rowCount": row_count,
        "ranking": ranking,
        "countries": countries,
        "sheets": sheets,
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({
        "output": str(output),
        "sheetCount": payload["sheetCount"],
        "rowCount": payload["rowCount"],
        "rankingRows": len(ranking),
        "countrySheets": len(countries),
    }, ensure_ascii=False))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python convert_benchmarking.py origem.xlsx saida.json")
    convert(sys.argv[1], sys.argv[2])
